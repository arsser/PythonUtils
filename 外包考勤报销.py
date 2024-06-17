#外包考勤报销，主要功能：
#1.从钉钉获取外包考勤、报销的审批记录；
#2.对审批记录做处理，包括：筛选上个月的数据，和hr的名单做merge
#3.按照供应商名单，将拆分后的文件生产邮件草稿

import requests
import csv
import time
import pandas as pd
from datetime import datetime,timedelta
import sys
import os
import win32com.client as win32
from openpyxl import load_workbook
import json
import glob

# 常量
APP_KEY = 'dingl5qlh2s1ddksf5ru'
APP_SECRET = 'mjr6mb1TpfMx1Q2W7IfSVvrTsD855sSMQcL_XNZQA_HVhCYBKT-FDm8VPqGpJxU4'
DINGTALK_API_BASE = 'https://oapi.dingtalk.com'
GET_TOKEN_URL = f'{DINGTALK_API_BASE}/gettoken?appkey={APP_KEY}&appsecret={APP_SECRET}'
APPROVAL_RECORD_URL = f'{DINGTALK_API_BASE}/topapi/processinstance/listids'

KQ_PROCESS_CODE = 'PROC-6B32AF04-FE29-4DAC-9B99-2D4D58183DC9' #外包考勤的OA审批,v1旧版版编辑模板时显示在url
BX_PROCESS_CODE = 'PROC-B193D2E7-4B1B-43EA-8581-CEF423B71C1F' #外包报销的OA审批

# 获取 AccessToken
def get_access_token():
    response = requests.get(GET_TOKEN_URL)
    if response.status_code == 200:
        return response.json()['access_token']
    else:
        raise Exception('Failed to get access token')

def datetime_to_timestamp(dt):
    return int(time.mktime(dt.timetuple()) * 1000)

# 获取审批记录 ID 列表
def get_all_approval_record_ids(token, process_code, start_time, end_time):
    all_record_ids = []
    size = 20
    next_cursor = 0
    #has_more = True
    # 在调用 API 之前，将 datetime 对象转换为时间戳
    start_timestamp = datetime_to_timestamp(start_time)
    end_timestamp = datetime_to_timestamp(end_time)
    while True:
        data = {
            'process_code': process_code,
            'start_time': start_timestamp,
            'end_time': end_timestamp,
            'size': size,
            'cursor': next_cursor
        }
        response = requests.post(f"{APPROVAL_RECORD_URL}?access_token={token}", json=data)
        
        response_data = response.json()

        # 检查 errcode
        if response_data['errcode'] != 0:
            # 打印或抛出异常信息
            error_msg = response_data.get('errmsg', 'Unknown error')
            raise Exception(f"获取{process_code}的审批列表 API Error: {error_msg}")
        
        if response.status_code == 200:
            result = response_data['result']
            all_record_ids.extend(result['list'])            
            if 'next_cursor' in result and result['next_cursor'] != 0:
                next_cursor = result['next_cursor']
            else:
                break  # 没有更多数据或next_cursor为0            
        else:
            raise Exception('Failed to get approval record IDs')
           
    return all_record_ids

# 获取单个审批记录的详细数据
def get_approval_record_details(token, record_id):
    details_url = f'{DINGTALK_API_BASE}/topapi/processinstance/get?access_token={token}'
    data = {'process_instance_id': record_id}
    response = requests.post(details_url, json=data)
    if response.status_code == 200:
        #print(response.json())
        return response.json()['process_instance']
    else:
        raise Exception('Failed to get approval record details')

# 分割标题并提取创建人的名字
def extract_creator_from_title(title):   
    parts = title.split('提交的')
    creator = parts[0] if len(parts) > 1 else None
    return creator

def prepare_data_for_excel(records):
    prepared_data = []
    # 遍历每条记录
    for record in records: 
        #form_values = {comp['name']: comp['value'] for comp in record['form_component_values']}

        # 解析表单组件的值, 在字典推导式中处理 'null' 值
        form_values = {
            comp['name']: ('' if comp['value'] == 'null' else comp['value']) 
            for comp in record['form_component_values']
        }

        if record['status'] == 'COMPLETED' and record['result'] == 'agree':
            # 整合其他字段
            combined_record = {
                '审批ID': record['business_id'],
                '审批结果': '审批通过',
                '审批状态': '已结束',
                '标题': record['title'],
                '创建人': extract_creator_from_title(record['title']),
                '创建人部门': record['originator_dept_name'],
                '发起时间': record['create_time'],
                '结束时间': record['finish_time'],
                **form_values  # 将表单值添加到记录中
            }
            prepared_data.append(combined_record)
    return prepared_data

#处理钉钉审批单中有嵌套表格的情况，并将表格明细按行展开
def parse_table_field_and_expand(record, table_data, form_values):
    expanded_records = []
    for row in table_data:
        parsed_row = {}
        for cell in row['rowValue']:
            if cell['componentType'] == 'DDAttachment':  # 特殊处理附件字段
                attachment_info = '; '.join([f"{att['fileName']} ({att['fileSize']} bytes)" for att in cell['value']])
                parsed_row[cell['label']] = attachment_info
            else:
                parsed_row[cell['label']] = cell['value']

        # 将表格的每一行作为一个独立的记录，包含主记录的其他信息
        expanded_record = {
            '审批ID': record['business_id'],
            '审批结果': '审批通过',
            '审批状态': '已结束',
            '标题': record['title'],
            '创建人': extract_creator_from_title(record['title']),
            '创建人部门': record['originator_dept_name'],
            '发起时间': record['create_time'],
            '结束时间': record['finish_time'],
            **form_values,  # 包含除表格外的其他表单值
            **parsed_row  # 包含当前表格行的信息
        }
        expanded_records.append(expanded_record)

    return expanded_records

def prepare_data_for_excel_forbx_expand(records):
    prepared_data = []
    for record in records:
        form_values = {comp['name']: comp['value'] for comp in record['form_component_values'] if comp['component_type'] != 'TableField'}

        if record['status'] == 'COMPLETED' and record['result'] == 'agree':
            # 对于包含表格的记录，单独处理
            for comp in record['form_component_values']:
                if comp['component_type'] == 'TableField':
                    table_data = json.loads(comp['value'])
                    prepared_data.extend(parse_table_field_and_expand(record, table_data, form_values))
                else:
                    combined_record = {
                        '审批ID': record['business_id'],
                        '审批结果': '审批通过',
                        '审批状态': '已结束',
                        '标题': record['title'],
                        '创建人': extract_creator_from_title(record['title']),
                        '创建人部门': record['originator_dept_name'],
                        '发起时间': record['create_time'],
                        '结束时间': record['finish_time'],
                        **form_values
                    }
                    prepared_data.append(combined_record)

    return prepared_data

#处理钉钉审批单中有嵌套表格的情况
# def parse_table_field_old(table_data):
    parsed_table = []
    for row in table_data:
        parsed_row = {}
        for cell in row['rowValue']:
            if cell['componentType'] == 'DDAttachment':  # 特殊处理附件字段
                attachment_info = '; '.join([f"{att['fileName']} ({att['fileSize']} bytes)" for att in cell['value']])
                parsed_row[cell['label']] = attachment_info
            else:
                parsed_row[cell['label']] = cell['value']
        parsed_table.append(parsed_row)
    return parsed_table

def parse_table_field(table_data):
    parsed_table = []
    for row in table_data:
        parsed_row = []
        for cell in row['rowValue']:
            if cell['componentType'] == 'DDAttachment':  # 特殊处理附件字段
                try:
                    attachment_info = '; '.join([f"{att['fileName']} ({att['fileSize']} bytes)" for att in cell['value']])
                    parsed_row.append(f"{cell['label']}: {attachment_info}")
                except Exception as e:
                    print(f"发生了一个意料之外的异常: {e}")
                    #raise  # 再次抛出异常，允许调试器捕捉到这个异常点
            else:
                parsed_row.append(f"{cell['label']}: {cell['value']}")
        parsed_table.append('\n'.join(parsed_row))
    return '\n'.join(parsed_table)  # 每个表格行之间加上换行符

#处理附件字段，把附件的文件名提取出来
# def extract_attachment_filenames_old1(attachment_data):
#     try:
#         # 将字符串数据解析为 JSON
#         attachments = json.loads(attachment_data)
#         # 提取所有文件名
#         file_names = [att['fileName'] for att in attachments]
#         return ', '.join(file_names)  # 用逗号分隔文件名
#     except json.JSONDecodeError:
#         # 如果解析出错，返回原始数据
#         return attachment_data

def extract_attachment_filenames(attachment_data):
    if not attachment_data:  # 检查数据是否为空或None
        return ''

    try:
        attachments = json.loads(attachment_data)
        if attachments is None:  # 检查解析结果是否为None
            return ''
        file_names = [att['fileName'] for att in attachments]
        return ', '.join(file_names)
    except json.JSONDecodeError:
        # 如果解析出错，返回原始数据或空字符串
        return ''

# 其余部分保持不变
def prepare_data_for_excel_forbx(records):
    prepared_data = []
    for record in records:
        form_values = {}
        for comp in record['form_component_values']:
            if comp['component_type'] == 'DDAttachment':
                # 特殊处理附件字段
                form_values[comp['name']] = extract_attachment_filenames(comp['value'])
            else:
                if comp['component_type'] == 'TableField':
                    # 解析表格
                    table_data = json.loads(comp['value'])
                    form_values[comp['name']] = parse_table_field(table_data)
                else:
                    value = comp['value']
                    if value == 'null':  # 将字符串'null'替换为空字符串
                        value = ''
                    form_values[comp['name']] = value                    

        if record['status'] == 'COMPLETED' and record['result'] == 'agree':
            combined_record = {
                '审批ID': record['business_id'],
                '审批结果': '审批通过',
                '审批状态': '已结束',
                '标题': record['title'],
                '创建人': extract_creator_from_title(record['title']),
                '创建人部门': record['originator_dept_name'],
                '发起时间': record['create_time'],
                '结束时间': record['finish_time'],
                **form_values
            }
            prepared_data.append(combined_record)
    return prepared_data

# def prepare_data_for_excel_forbx_old2(records):
    prepared_data = []
    for record in records:
        form_values = {}
        for comp in record['form_component_values']:
            if comp['component_type'] == 'TableField':
                # 解析表格
                table_data = json.loads(comp['value'])
                form_values[comp['name']] = parse_table_field(table_data)
            else:
                form_values[comp['name']] = comp['value']

        if record['status'] == 'COMPLETED' and record['result'] == 'agree':
            combined_record = {
                '审批ID': record['business_id'],
                '审批结果': '审批通过',
                '审批状态': '已结束',
                '标题': record['title'],
                '创建人': extract_creator_from_title(record['title']),
                '创建人部门': record['originator_dept_name'],
                '发起时间': record['create_time'],
                '结束时间': record['finish_time'],
                **form_values
            }
            prepared_data.append(combined_record)
    return prepared_data

# def prepare_data_for_excel_forbx_old1(records):
    prepared_data = []

    for record in records:
        # 解析表单组件的值
        form_values = {}
        for comp in record['form_component_values']:
            if comp['component_type'] == 'TableField':  # 假设表格的类型是 'TableField'
                table_data = []
                for row in comp['value']:  # 假设每行数据是一个字典
                    row_data = {cell['label']: cell['value'] for cell in row}  # 提取每行的单元格数据
                    table_data.append(row_data)
                form_values[comp['name']] = table_data
            else:
                form_values[comp['name']] = comp['value']

        if record['status'] == 'COMPLETED' and record['result'] == 'agree':
            combined_record = {
                '审批ID': record['business_id'],
                '审批结果': '审批通过',
                '审批状态': '已结束',
                '标题': record['title'],
                '创建人': extract_creator_from_title(record['title']),
                '创建人部门': record['originator_dept_name'],
                '发起时间': record['create_time'],
                '结束时间': record['finish_time'],
                **form_values
            }
            prepared_data.append(combined_record)

    return prepared_data

def get_approval_data(token, process_code, start_time, end_time):
    record_ids = get_all_approval_record_ids(token, process_code, start_time, end_time)
    detailed_records = [get_approval_record_details(token, record_id) for record_id in record_ids]
    return detailed_records

def number_to_chinese(num):
    num_to_chinese = {
        1: '一',
        2: '二',
        3: '三',
        4: '四',
        5: '五',
        6: '六',
        7: '七',
        8: '八',
        9: '九',
        10: '十',
        11: '十一',
        12: '十二'
    }
    return num_to_chinese.get(num, '')

#获取考勤的钉钉审批记录，kaoqin_month表示要获取哪个月份的考勤记录;
def get_kaoqin_ding_data(token, process_code,  start_time, end_time, kaoqin_month, excel_file):
    record_ids = get_all_approval_record_ids(token, process_code, start_time, end_time)
    print(f'{start_time} - {end_time} 考勤审批记录条数：{len(record_ids)}')
    detailed_records = [get_approval_record_details(token, record_id) for record_id in record_ids]
    prepared_records =  pd.DataFrame(prepare_data_for_excel(detailed_records))
    # 过滤出 '出勤月份' 为 kaoqin_month 的数据
    kaoqin_mon_chin = f'{number_to_chinese(kaoqin_month)}月'
    df_filtered = prepared_records[prepared_records['出勤月份'] == kaoqin_mon_chin]
    print(f'{kaoqin_month}月考勤记录条数：{len(df_filtered)}')
    df_filtered.to_excel(excel_file, index=False)
    print(f'保存{kaoqin_month}月考勤记录到{excel_file}')
    return 0

def get_baoxiao_ding_data(token, process_code,  start_time, end_time, baoxiao_month, excel_file):
    record_ids = get_all_approval_record_ids(token, process_code, start_time, end_time)
    print(f'获取报销审批记录条数：{len(record_ids)}')
    detailed_records = [get_approval_record_details(token, record_id) for record_id in record_ids]
    prepared_records =  pd.DataFrame(prepare_data_for_excel_forbx(detailed_records))    
    df_filtered = prepared_records  
    df_filtered.to_excel(excel_file, index=False)
    print(f'保存{baoxiao_month}月报销记录到{excel_file}')
    return 0

#把考勤的excel和hr的人员信息excel进行合并,该excel按月份建立sheet，记录当月的外包人员信息；
def merge_kaoqin_hr_excel(kaoqin_excel, hr_excel, hr_excel_sheet, merged_excel):
    # 读取审批记录表
    df_approval_records = pd.read_excel(kaoqin_excel)

    # 读取公司信息表的上个月数据
    df_company_info = pd.read_excel(hr_excel, sheet_name=hr_excel_sheet,usecols=['姓名','供应商','入职日期','离职日期'])
     
    # 使用 '姓名'（或其他合适的键）进行合并
    # 假设两个表都有一个共同的列 '姓名' 用于合并
    df_merged = pd.merge(df_approval_records, df_company_info, left_on='创建人', right_on='姓名', how='outer')
    df_merged['入职日期'] = pd.to_datetime(df_merged['入职日期']).dt.date
    df_merged['离职日期'] = pd.to_datetime(df_merged['离职日期']).dt.date
    # 删除重复的列（例如，删除 "姓名" 列）
    #df_merged.drop(columns=['姓名'], inplace=True)
    # 指定新的列顺序
    new_column_order = ['供应商', '审批ID', '姓名','出勤月份', '入职日期','离职日期','创建人',
                        '本月缺勤天数','具体缺勤日期','本月加班天数','具体加班日期',
                        '加班原因','备注','审批状态','审批结果','标题','发起时间','结束时间',
                        '创建人部门','部门信息']

    # 重新排列列
    df_merged = df_merged[new_column_order]
    # 将合并后的数据导出到新的 Excel 文件
    df_merged.to_excel(merged_excel, index=False)
    print(f'merge考勤记录到{merged_excel}')

#把报销的excel和hr的人员信息excel进行合并
def merge_baoxiao_hr_excel(baoxiao_excel, hr_excel, hr_excel_sheet, merged_excel):
    # 读取审批记录表
    df_approval_records = pd.read_excel(baoxiao_excel)

    # 读取公司信息表的上个月数据
    df_company_info = pd.read_excel(hr_excel, sheet_name=hr_excel_sheet,usecols=['姓名','供应商','入职日期','离职日期'])
     
    # 使用 '姓名'（或其他合适的键）进行合并
    # 假设两个表都有一个共同的列 '姓名' 用于合并
    df_merged = pd.merge(df_approval_records, df_company_info, left_on='创建人', right_on='姓名', how='outer')
    df_merged['入职日期'] = pd.to_datetime(df_merged['入职日期']).dt.date
    df_merged['离职日期'] = pd.to_datetime(df_merged['离职日期']).dt.date
    # 删除重复的列（例如，删除 "姓名" 列）
    #df_merged.drop(columns=['姓名'], inplace=True)
    # 指定新的列顺序
    new_column_order = ['供应商', '审批ID', '姓名','报销归属月份', '入职日期','离职日期', '创建人',                       
                        '报销总金额（元）','表格','其他附件','备注','审批状态',
                        '审批结果','标题','发起时间','结束时间','创建人部门']
    # 重新排列列
    df_merged = df_merged[new_column_order]
    # 将合并后的数据导出到新的 Excel 文件
    df_merged.to_excel(merged_excel, index=False)
    print(f'merge报销记录到{merged_excel}')

#按供应商把考勤excel文件做拆分
def split_excel(excel, month, type):
    df = pd.read_excel(excel)
    # 如果存在日期列，转换为仅日期格式
    if '入职日期' in df.columns:
        df['入职日期'] = pd.to_datetime(df['入职日期']).dt.date
    if '离职日期' in df.columns:
        df['离职日期'] = pd.to_datetime(df['离职日期']).dt.date
    for value in df['供应商'].unique():
        df_subset = df[df['供应商'] == value]
        base_dir = os.path.dirname(excel)
        df_subset.to_excel(rf"{base_dir}\chai_{value}_{month}月{type}.xlsx", index=False)
    print(f"按供应商分拆{type}excel完成，共{len(df_subset)}个文件")

def mailto_supplier(account_email, supplier_info, directory, subject_suffix_year_month):
    outlook = win32.Dispatch('outlook.application')
    namespace = outlook.GetNamespace("MAPI")
    account = namespace.Folders[account_email]

    # 获取当前日期和前一个月的日期
    current_date = datetime.now()
    previous_month_date = current_date.replace(day=1) - timedelta(days=1)

    # 格式化日期为 "2023年11月" 和 "2023年x月x日"
    subject_suffix = subject_suffix_year_month #previous_month_date.strftime(" - %Y年%m月")
    closing_remark = f"依图供应链\n{current_date.strftime('%Y年%m月%d日')}"

    for supplier in supplier_info:
        # 查找并附加文件（使用模糊匹配）
        files = find_files(supplier['供应商'], directory)

        # 如果没有找到文件，则跳过当前供应商
        if not files:
            print(f"未找到与供应商 {supplier['供应商']} 匹配的excel文件，跳过发邮件")
            continue
        
        mail = account.Items.Add("IPM.Note")
        mail.To = supplier['邮箱']
        mail.Subject = f"请确认软件技术服务人员出勤-报销{subject_suffix}"
        mail.Body = f"尊敬的供应商，{supplier['供应商']}，\n\n相关内容，请查收附件。\n\n{closing_remark}"
        for file in files:
            mail.Attachments.Add(file)

        mail.Save()
        print(f"邮件草稿已保存：{supplier['供应商']}")

def find_files(supplier_name_fragment, directory):
    """根据供应商名称的一部分查找文件"""
    files = []
    for file in os.listdir(directory):
        if supplier_name_fragment.lower() in file.lower() and file.endswith(".xlsx"):
            files.append(os.path.join(directory, file))
    return files

def read_excel(file_path):
    """从 Excel 读取供应商信息"""
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    suppliers = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        suppliers.append({'供应商': row[0], '联系人': row[1], '邮箱': row[2]})
    return suppliers

# 主逻辑
if __name__ == '__main__':     
    # 获取当前日期和时间，并格式化为字符串，格式为"年-月-日-时分秒"
    current_datetime_str = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    base_dir = fr'\\10.12.21.65\share\外包费用\新流程'
    kaoqin_excel = fr'{base_dir}\考勤报销数据\外包出勤审批_{current_datetime_str}.xlsx'
    kaoqin_merged_excel = fr'{base_dir}\考勤报销数据\外包出勤审批_merge_{current_datetime_str}.xlsx'
    hr_excel =           fr'{base_dir}\【HR】外包人员表.xlsx'
    
    baoxiao_excel = fr'{base_dir}\考勤报销数据\外包报销审批_{current_datetime_str}.xlsx'
    
    baoxiao_merged_excel = fr'{base_dir}\考勤报销数据\外包报销审批_merge_{current_datetime_str}.xlsx'

    #debug
    #kaoqin_excel_file = f'\\10.12.21.65\share\外包费用\新流程\考勤报销数据\外包报销审批_2024-03-12-181914.xlsx'
    #baoxiao_excel = fr'\\10.12.21.65\share\外包费用\新流程\考勤报销数据\外包出勤审批_2024-03-12-181914.xlsx'
    #kaoqin_excel=fr'{base_dir}\考勤报销数据\外包出勤审批_2024-03-12-181914.xlsx'
    #baoxiao_excel=fr'{base_dir}\考勤报销数据\外包报销审批_2024-03-12-181914.xlsx'
    #kaoqin_merged_excel = fr'{base_dir}\考勤报销数据\外包出勤审批_merge_2024-02-07-122825.xlsx'
    #baoxiao_merged_excel = fr'{base_dir}\考勤报销数据\外包报销审批_merge_2024-02-07-122825.xlsx'
    #debug

    token = get_access_token()

    # 获取当前时间
    now = datetime.now()
    #考勤的数据拉取时间从上个月的1号到今天;
    kq_data_end_time = now 
    # 计算上个月的第一天, 如果当前是1月，则上个月是去年的12月
    if now.month == 1:
        kq_data_start_time = datetime(now.year - 1, 12, 1)
    else:
        kq_data_start_time = datetime(now.year, now.month - 1, 1)
    
    #报销的数据拉取时间从上个月的1号到上个月的最后一天;
    bx_data_end_time = datetime(now.year, now.month, 1)
    if now.month == 1:
        bx_data_start_time = datetime(now.year - 1, 12, 1)
    else:
        bx_data_start_time = datetime(now.year, now.month - 1, 1)

    #0. 如果文件已存在，就跳过1-3步，直接生成邮件草稿；
    # 指定要搜索的路径
    path = fr'{base_dir}\考勤报销数据'
    pattern = '*.xlsx'
    # 使用glob.glob()搜索匹配的文件
     # 获取上一个月的月份
    now = datetime.now()
    first_day_of_current_month = datetime(now.year, now.month, 1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    previous_month_str = last_day_of_previous_month.strftime("%Y-%m")
    matching_files = glob.glob(os.path.join(path, pattern))
    if not matching_files:
        #1. 获取考勤,报销的审批记录，并保存为excel文件；
        print("#1. 获取考勤,报销的审批记录，并保存为excel文件；")
        get_kaoqin_ding_data(token, KQ_PROCESS_CODE, kq_data_start_time, kq_data_end_time, kq_data_start_time.month, kaoqin_excel)
        get_baoxiao_ding_data(token, BX_PROCESS_CODE, bx_data_start_time, bx_data_end_time, bx_data_start_time.month, baoxiao_excel)
        #sys.exit(0)
    else:
        print("#1.本地目录找到文件，跳过钉钉；")
    #2. merge       
    merge_kaoqin_hr_excel(kaoqin_excel, hr_excel, previous_month_str, kaoqin_merged_excel)
    merge_baoxiao_hr_excel(baoxiao_excel, hr_excel, previous_month_str, baoxiao_merged_excel)
    #sys.exit(0)
    
    #3. 分拆    
    split_excel(kaoqin_merged_excel, previous_month_str, "出勤")
    split_excel(baoxiao_merged_excel, previous_month_str, "报销")
    #sys.exit(0)

    #4. 生成邮件草稿
    outlook_account = 'scm_bill@yitu-inc.com' # 替换为你的 Outlook 账户名
    directory = base_dir 
    suppliers = read_excel(base_dir +"\人力外包供应商信息.xlsx")
    mailto_supplier(outlook_account, suppliers, base_dir+'\考勤报销数据',previous_month_str)

