#1.把供应商邮件回复的excel结算文件从outlook指定拷贝到指定的目录，并进行合并，如果文件已存在不覆盖（因为可能修改过了）
#  邮件判别规则：类别为“待AI处理结算单"，附件判别规则：主题含有关键字："结算单"；
#  
#2.加入把HR的数据人员基本薪资，并计算二者的差额，进行对账，差额不为0的高亮显示；
#3.

import os
import pandas as pd
import win32com.client
from openpyxl import load_workbook
from datetime import datetime, timedelta
from datetime import date

# 设置 Outlook 连接
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# 选择文件夹（例如收件箱）
inbox = outlook.GetDefaultFolder(6)  # 6 代表收件箱

# 搜索符合条件的邮件
messages = inbox.Items
messages = messages.Restrict("[Categories] = '待AI处理结算单'")

# 指定文件保存的目录
save_folder = fr'\\10.12.21.65\share\外包费用\新流程\人力供应商结算单'
merged_file = fr'\\10.12.21.65\share\外包费用\新流程\人力供应商结算单-merged.xlsx'
template_file = fr'\\10.12.21.65\share\外包费用\新流程\软件技术服务人员结算单(231205版)-合集.xlsx'

def get_account(folders, account_email):
    """
    从 Outlook 的文件夹列表中选择与指定电子邮件地址对应的账户。

    :param folders: Outlook 文件夹列表
    :param account_email: 要选择的账户的电子邮件地址
    :return: 选定的账户文件夹
    """
    for folder in folders:
        try:
            if folder.Store and  folder.Store.DisplayName.lower() == account_email.lower():
                return folder
        except AttributeError:
            continue
    return None

def download_attachments(category, subject_keyword, save_folder):
    """
    下载符合条件的邮件中的 Excel 附件。

    :param folder_name: 邮件文件夹名称（如 "Inbox"）
    :param category: 邮件类别
    :param subject_keyword: 邮件主题关键词
    :param save_folder: 附件保存目录
    """
    folder_name = "收件箱"
    try:
        # 选择账户
        account_name = "scm_bill@yitu-inc.com"
        account = get_account(outlook.Folders, account_name)
        if not account:
            print("账户未找到")
            exit()
        folder = account.Folders[folder_name]
        messages = folder.Items
        query_string = "[Categories] = '{}'".format(category)
        #print(query_string)
        messages = messages.Restrict(query_string)        
        for message in messages:
            #print("正在处理邮件：{}".format(message.Subject))
            if not subject_keyword or subject_keyword in message.Subject:
                attachments = message.Attachments
                for attachment in attachments:
                    if '结算单' in attachment.FileName:
                        file_path = os.path.join(save_folder, attachment.FileName)
                        if not os.path.exists(file_path):  # 检查文件是否存在
                            attachment.SaveAsFile(file_path)
                            print(f"保存附件：{file_path}")
                        else:
                            print(f"文件已存在，跳过保存：{file_path}")
                # 清除邮件的类别标记
                #message.Categories = ''
                #message.Save()
    except Exception as e:
        print("Error while downloading attachments: ", e)


def merge_excel_files(sheet_name1 ,sheet_name2,  directory, file_extension=".xlsx", ):
    all_data = pd.DataFrame()
    for file in os.listdir(directory):
        if file.endswith(file_extension):
            file_path = os.path.join(directory, file)
            print("正在处理文件：{}".format(file))

            # 使用 ExcelFile 加载文件，以便检查工作表名称
            xls = pd.ExcelFile(file_path)            
            # 检查 sheet_name1 是否存在
            if sheet_name1 in xls.sheet_names: 
                sheet_name1 = sheet_name1
            else:
                sheet_name1 = sheet_name2

            # 指定读取的列和跳过的行数
            #df = pd.read_excel(file_path, usecols='C:F,O,Q:T', skiprows=2)
            if (sheet_name1):
                df = pd.read_excel(file_path, sheet_name=sheet_name1, skiprows=2)
            else:
                df = pd.read_excel(file_path, skiprows=2)
            #print(df.head())
            # 找到C列中最后一个有值的行的索引            
            last_valid_index = df.iloc[:, 2].last_valid_index()
            # 如果找到有效行，截取到这一行
            if last_valid_index is not None:
                df = df.loc[:last_valid_index]
            all_data = pd.concat([all_data, df], ignore_index=True)    
    print(all_data)        
    return all_data

def merge_excel_files2(sheet_name1, sheet_name2, directory, file_extension=".xlsx", columns=None):
    all_data = pd.DataFrame()
    for file in os.listdir(directory):
        if file.endswith(file_extension):
            file_path = os.path.join(directory, file)
            print("正在处理文件：{}".format(file))

            xls = pd.ExcelFile(file_path)
            # 使用sheet_name2如果sheet_name1不存在
            actual_sheet_name = sheet_name1 if sheet_name1 in xls.sheet_names else sheet_name2

            # 读取数据，应用usecols限制导入的列
            df = pd.read_excel(file_path, sheet_name=actual_sheet_name, skiprows=2, usecols=columns)

            last_valid_index = df.iloc[:, 2].last_valid_index()
            if last_valid_index is not None:
                df = df.loc[:last_valid_index]
            all_data = pd.concat([all_data, df], ignore_index=True)

    print(all_data)
    return all_data


def copy_data_to_template(data, template_path, start_row, cols_to_copy, sheet_name):
    """
    将 DataFrame 数据拷贝到 Excel 模板文件的指定列和指定工作表中，并直接保存在模板路径。
    :param data: 要拷贝的数据（DataFrame）
    :param template_path: 模板文件路径
    :param start_row: 开始拷贝的起始行
    :param cols_to_copy: 要拷贝的列的映射（字典），格式为 {'数据列名': '模板列名'}
    :param sheet_name: 要拷贝到的工作表名称
    """
    workbook = load_workbook(template_path)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    worksheet = workbook[sheet_name]

    for i, row in data.iterrows():
        for data_col, template_col in cols_to_copy.items():
            cell = worksheet[f'{template_col}{start_row + i}']
            cell.value = row[data_col]

    workbook.save(template_path)

def convert_indexes_to_column_names(dataframe, indexes):
    """
    将 DataFrame 中的列索引转换为列名。

    :param dataframe: 数据源 DataFrame
    :param indexes: 要转换的列索引列表
    :return: 转换后的列名列表
    """
    column_names = []
    for index in indexes:
        try:
            column_names.append(dataframe.columns[index])
        except IndexError:
            # 如果索引超出范围，可以选择抛出错误或者添加错误处理逻辑
            print(f"Index {index} is out of range.")
    return column_names

# 使用这些函数的示例代码
if __name__ == "__main__":
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    #print(pd.__version__)
    # 设置参数
    #save_folder = "C:\\path\\to\\your\\directory"
    
    category = "待AI处理结算单"
    subject_keyword = ""
    
    # 下载附件
    #debug download_attachments(category, subject_keyword, save_folder)

    # 合并 Excel 文件
    last_month = datetime.now() - timedelta(days=45)
    #last_month = date(2024, 3, 1)
    sheet_name1 = last_month.strftime("结算单%Y-%m")
    sheet_name2 = sheet_name1.replace("-0","-")
    merged_data = merge_excel_files2(sheet_name1, sheet_name2, save_folder, ".xlsx", "A:AC")

    # 保存合并后的文件
    merged_data.to_excel(os.path.join(save_folder, merged_file), index=False)

    #exit(0)

    data_file = merged_file  # 数据文件路径
    template_path =  template_file
    #无用output_path = fr'\\10.12.21.65\share\外包费用\新流程\[整合].xlsx'
    start_row = 4  # 从模板的第4行开始拷贝数据
   
    # 从文件读取数据
    data = pd.read_excel(data_file)

    # 定义要拷贝的列的序号和模板中对应的列名
    indexes_to_convert = [1,  2,  3,  4,  5,  14, 16, 17, 18, 19, 20, 21, 22, 23,  24,  25,  26, 27,  28]  # 数据列的序号，例如第1、2、3列
    template_columns =   ['B','C','D','E','H','Q','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD',"AE"]  # 模板中对应的列名

    # 将序号转换为列名
    converted_column_names = convert_indexes_to_column_names(data, indexes_to_convert)

    # 创建列的映射关系
    cols_to_copy = dict(zip(converted_column_names, template_columns))

    
    # 调用函数将数据拷贝到模板
    # 计算上个月的年-月格式
    #无用last_month = datetime.now() - timedelta(days=30)
    sheet_name = "结算单"+last_month.strftime("%Y-%m")
    
    copy_data_to_template(data, template_path, start_row, cols_to_copy, sheet_name)
    print("Completed!")