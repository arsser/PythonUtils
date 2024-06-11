import os
from datetime import datetime, timedelta
import win32com.client as win32
from openpyxl import load_workbook

def find_files_1(supplier_name, directory):
    """根据供应商名称查找文件"""
    files = []
    for file in os.listdir(directory):
        if file.startswith(f"chai_{supplier_name}_") and file.endswith(".xlsx"):
            files.append(os.path.join(directory, file))
    return files

def find_files(supplier_name_fragment, directory):
    """根据供应商名称的一部分查找文件"""
    files = []
    for file in os.listdir(directory):
        if supplier_name_fragment.lower() in file.lower() and file.endswith(".xlsx"):
            files.append(os.path.join(directory, file))
    return files

def read_excel(file_path):
    """从 Excel 读取供应商信息"""
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    suppliers = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        suppliers.append({'供应商': row[0], '联系人': row[1], '邮箱': row[2]})
    return suppliers



def save_draft_email_1(account_email, supplier_info, directory):
    outlook = win32.Dispatch('outlook.application')
    namespace = outlook.GetNamespace("MAPI")
    account = namespace.Folders[account_email]

    for supplier in supplier_info:
        mail = account.Items.Add("IPM.Note")
        mail.To = supplier['邮箱']
        mail.Subject = f"请确认软件技术服务人员出勤-报销"
        mail.Body = f"{supplier['供应商']}，相关内容，请查收附件。"

        # 查找并附加文件
        files = find_files(supplier['供应商'], directory)
        # 如果没有找到文件，则跳过当前供应商
        if not files:
            print(f"未找到与供应商 {supplier['供应商']} 匹配的文件，跳过发送邮件。")
            continue

        for file in files:
            mail.Attachments.Add(file)
        mail.Save()


def save_draft_email(account_email, supplier_info, directory):
    outlook = win32.Dispatch('outlook.application')
    namespace = outlook.GetNamespace("MAPI")
    account = namespace.Folders[account_email]

    # 获取当前日期和前一个月的日期
    current_date = datetime.now()
    previous_month_date = current_date.replace(day=1) - timedelta(days=1)

    # 格式化日期为 "2023年11月" 和 "2023年x月x日"
    subject_suffix = previous_month_date.strftime(" - %Y年%m月")
    closing_remark = f"依图供应链\n{current_date.strftime('%Y年%m月%d日')}"

    for supplier in supplier_info:
        # 查找并附加文件（使用模糊匹配）
        files = find_files(supplier['供应商'], directory)

        # 如果没有找到文件，则跳过当前供应商
        if not files:
            print(f"未找到与供应商 {supplier['供应商']} 匹配的文件，跳过发送邮件。")
            continue

        mail = account.Items.Add("IPM.Note")
        mail.To = supplier['邮箱']
        mail.Subject = f"请确认软件技术服务人员出勤-报销{subject_suffix}"
        mail.Body = f"尊敬的供应商，{supplier['供应商']}，\n\n相关内容，请查收附件。\n\n{closing_remark}"

        for file in files:
            mail.Attachments.Add(file)

        mail.Save()


# 脚本主体
excel_file_path = fr"\\10.12.21.65\Share\外包费用\新流程\人力外包供应商信息.xlsx" # 替换为你的 Excel 文件路径
outlook_account = 'scm_bill@yitu-inc.com' # 替换为你的 Outlook 账户名
directory = fr'\\10.12.21.65\Share\外包费用\新流程\2023-11\Step1-考勤数据' # 替换为文件所在的目录

suppliers = read_excel(excel_file_path)
save_draft_email(outlook_account, suppliers, directory)
